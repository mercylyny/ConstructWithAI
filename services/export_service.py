import os
import csv
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, Any

# Ensure output directory exists
EXPORT_DIR = "outputs/exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

def export_boq_to_csv(boq_data: Dict[str, Any]) -> str:
    """
    Exports BOQ data to a CSV file.
    Returns the absolute file path.
    """
    filename = f"boq_export_{uuid.uuid4().hex[:8]}.csv"
    filepath = os.path.join(EXPORT_DIR, filename)

    line_items = boq_data.get("line_items", [])
    total_cost = boq_data.get("total_project_cost", 0.0)
    currency = boq_data.get("currency", "UGX")

    with open(filepath, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        
        # Header
        writer.writerow(["Item", "Quantity", "Unit", f"Rate ({currency})", f"Amount ({currency})"])
        
        # Rows
        for item in line_items:
            writer.writerow([
                item.get("item"),
                item.get("quantity"),
                item.get("unit"),
                item.get("rate"),
                item.get("amount")
            ])
            
        # Total Row
        writer.writerow([])
        writer.writerow(["TOTAL PROJECT COST", "", "", "", total_cost])
        
    return os.path.abspath(filepath)

def export_boq_to_excel(boq_data: Dict[str, Any]) -> str:
    """
    Exports BOQ data to an Excel file using openpyxl.
    Returns the absolute file path.
    """
    filename = f"boq_export_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    
    line_items = boq_data.get("line_items", [])
    total_cost = boq_data.get("total_project_cost", 0.0)
    currency = boq_data.get("currency", "UGX")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"
    
    # Headers
    headers = ["Item", "Quantity", "Unit", f"Rate ({currency})", f"Amount ({currency})"]
    ws.append(headers)
    
    # Style Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Rows
    for item in line_items:
        row = [
            item.get("item"),
            item.get("quantity"),
            item.get("unit"),
            item.get("rate"),
            item.get("amount")
        ]
        ws.append(row)
        
    # Total Row
    total_row_idx = len(line_items) + 3 # Header + Data + Spacer
    ws.cell(row=total_row_idx, column=1, value="TOTAL PROJECT COST").font = Font(bold=True)
    ws.cell(row=total_row_idx, column=5, value=total_cost).font = Font(bold=True)
    
    # Auto-adjust column widths (simple approximation)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    wb.save(filepath)
    return os.path.abspath(filepath)
